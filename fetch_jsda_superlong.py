#!/usr/bin/env python3.12
"""Fetch JSDA super-long JGB net-buying by investor and build superlong_data.json.

Source: JSDA 公社債投資家別売買高 (incl. 国債投資家別売買高), per-fiscal-year workbooks
under .../toukei/tentoubaibai/. We read sheet "(Ｋ)一般差引" (outright net, all-member
basis, ex-repo), column 超長期 (JGBs over 10-year), for the target investor rows.

Sign: the sheet reports 差引 = 売付額 − 買付額 (Sell − Buy). Net BUYING = −(差引).
Output values are in ¥tn (億円 / 10000), positive = net buying of super-long JGBs.

Seasonal layout: one line per Japanese fiscal year (FY = Apr..Mar), keyed by FY start year.

Run:  python3.12 fetch_jsda_superlong.py
"""
import urllib.request, socket, time, json, os, datetime

BASE = "https://www.jsda.or.jp/shiryoshitsu/toukei/tentoubaibai/"
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source"); os.makedirs(SRC, exist_ok=True)

# fiscal-year -> workbook filename on JSDA (r suffix varies by year)
FY_FILES = {
    2021: "koushasai2021r.xlsx",
    2022: "koushasai2022r.xlsx",
    2023: "koushasai2023.xlsx",
    2024: "koushasai2024.xlsx",
    2025: "koushasai2025.xlsx",
    2026: "koushasai.xlsx",          # current fiscal year (partial)
}
# investor JP row label -> output key + display label
TARGETS = {
    "生保・損保":  ("life",    "Life & non-life insurers"),
    "信託銀行":    ("trust",   "Trust banks"),
    "外国人":      ("foreign", "Foreigners"),
    "都市銀行":    ("mega",    "City (mega) banks"),
}
MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]

# ── "by player" flow chart (matches Macrobond "Japan Bond Flows") ──────────────
# Named investor buckets (matched by JP-label prefix); everything else folds into
# an "others" residual so each month reconciles exactly to the 合計 (Total) row.
# order = stacking order; the residual "others" is appended last in the chart.
PLAYERS = [
    # jp-prefix        key         label                        color
    ("都市銀行",       "mega",     "City (mega) banks",         "#4cba7a"),
    ("地方銀行",       "regional", "Regional banks",            "#3fb8a0"),
    ("信託銀行",       "trust",    "Trust banks (pension)",     "#4a90d9"),
    ("農林系金融機関", "agri",     "Agri co-ops (Norinchukin)", "#b79ae8"),
    ("生保・損保",     "life",     "Life & non-life insurers",  "#d9a441"),
    ("投資信託",       "invtrust", "Investment trusts",         "#e08fb0"),
    ("外国人",         "foreign",  "Foreigners",                "#e2564a"),
    ("債券ディーラー", "dealer",   "Bond dealers",              "#8a94ad"),
    ("その他",         "othcat",   "Other (JSDA その他)",       "#cf8a5a"),  # exact-match; big dealer-offset line
]
OTHERS = ("others", "Others (minor)", "#55617a")

def get(url, timeout=90, retries=4, sleep=8):
    last = None
    for i in range(retries):
        try:
            with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last = e; print(f"  retry {i+1}/{retries} {url}: {e}"); time.sleep(sleep*(i+1))
    raise last

def norm(x):
    return str(x).replace("_x000D_","").replace("\n","").replace("\r","").strip() if x is not None else ""

def parse_fy(path, fy):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "一般差引" in s and "証券" not in s), None)
    if not sheet:
        raise RuntimeError(f"no 一般差引 sheet in {path}")
    sh = wb[sheet]
    rows = [list(r) for r in sh.iter_rows(values_only=True)]
    # locate the 超長期 column from the header band (first ~6 rows)
    cho = None
    for r in rows[:6]:
        for i, c in enumerate(r):
            if norm(c).startswith("超長期"):
                cho = i; break
        if cho is not None: break
    if cho is None:
        raise RuntimeError(f"no 超長期 column in {path}")
    out = {v[0]: [None]*12 for v in TARGETS.values()}
    for r in rows:
        ym, inv = norm(r[0]), norm(r[1])
        if inv in TARGETS and "/" in ym and isinstance(r[cho], (int, float)):
            y, m = ym.split("/"); m = int(m)
            idx = (m - 4) % 12                    # Apr->0 ... Mar->11
            key = TARGETS[inv][0]
            out[key][idx] = round(-r[cho] / 10000.0, 3)   # net buying, ¥tn
    return out

def parse_players(path):
    """Per-investor monthly net buying, two measures: all-JGBs-ex-T-bills and
    super-long (>10Y). Returns {month_tag -> {playerkey|'total' -> {'ext','sl'}}}.
    month_tag is 'YYYY-MM'. Net buying = purchases − sales = −(差引), ¥tn."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "一般差引" in s and "証券" not in s), None)
    if not sheet:
        raise RuntimeError(f"no 一般差引 sheet in {path}")
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    # locate columns: 国債 total (jgb), 超長期 (sl), 国庫短期証券等 (tbills)
    jgb = sl = tb = None
    for r in rows[:6]:
        for i, c in enumerate(r):
            v = norm(c)
            if v.startswith("超長期") and sl is None: sl = i
            if "国庫短期証券" in v and tb is None: tb = i
            if v.startswith("国債") and "Government" in v and jgb is None: jgb = i
    if None in (jgb, sl, tb):
        raise RuntimeError(f"missing JGB columns in {path} (jgb={jgb},sl={sl},tb={tb})")
    def playerkey(inv):
        if inv == "その他": return "othcat"           # exact — not その他金融機関/その他法人
        for pre, key, *_ in PLAYERS:
            if key == "othcat": continue               # handled above by exact match
            if inv.startswith(pre): return key
        return OTHERS[0]
    out = {}
    for r in rows:
        ym, inv = norm(r[0]), norm(r[1])
        parts = ym.split("/")
        if len(parts) < 2 or not inv: continue
        try: y, m = int(parts[0]), int(parts[1])
        except ValueError: continue
        tag = f"{y}-{m:02d}"
        jv, sv, tv = r[jgb], r[sl], r[tb]
        if not isinstance(jv, (int, float)): continue
        ext = -(jv - (tv if isinstance(tv, (int, float)) else 0)) / 10000.0
        slv = -(sv / 10000.0) if isinstance(sv, (int, float)) else None
        bucket = "total" if inv.startswith("合計") else playerkey(inv)
        d = out.setdefault(tag, {})
        cur = d.setdefault(bucket, {"ext": 0.0, "sl": 0.0})
        cur["ext"] += ext
        if slv is not None: cur["sl"] += slv
    return out

def build_players(all_months):
    """all_months: dict month_tag -> {bucket -> {'ext','sl'}} merged across FYs.
    Produces chronological arrays per player + total + others residual."""
    months = sorted(all_months)
    keys = [p[1] for p in PLAYERS]
    ext = {k: [] for k in keys + ["others", "total"]}
    slg = {k: [] for k in keys + ["others", "total"]}
    for t in months:
        d = all_months[t]
        tot_e = d.get("total", {}).get("ext")
        tot_s = d.get("total", {}).get("sl")
        named_e = named_s = 0.0
        for k in keys:
            e = d.get(k, {}).get("ext", 0.0); s = d.get(k, {}).get("sl", 0.0)
            ext[k].append(round(e, 3)); slg[k].append(round(s, 3))
            named_e += e; named_s += s
        # residual others = total − named (keeps the stack summing to 合計)
        ext["others"].append(round((tot_e - named_e), 3) if tot_e is not None else None)
        slg["others"].append(round((tot_s - named_s), 3) if tot_s is not None else None)
        ext["total"].append(round(tot_e, 3) if tot_e is not None else None)
        slg["total"].append(round(tot_s, 3) if tot_s is not None else None)
    order = keys + ["others"]
    labels = {p[1]: p[2] for p in PLAYERS}; labels["others"] = OTHERS[1]
    colors = {p[1]: p[3] for p in PLAYERS}; colors["others"] = OTHERS[2]
    return {"months": months, "order": order, "labels": labels, "colors": colors,
            "extbills": ext, "superlong": slg}

def main():
    socket.setdefaulttimeout(120)
    series = {k: {"label": lbl, "fy": {}} for k, (kk, lbl) in
              [(v[0], v) for v in TARGETS.values()]}
    latest_month = None
    players_raw = {}
    for fy, fname in FY_FILES.items():
        path = os.path.join(SRC, fname)
        try:
            # Historical fiscal-year workbooks are static once published — reuse the
            # cached copy. Only the current-year file (koushasai.xlsx) is refetched
            # each run. This also sidesteps JSDA throttling on bulk downloads.
            cached = os.path.exists(path) and os.path.getsize(path) > 3000
            if cached and fname != "koushasai.xlsx":
                print(f"FY{fy}: using cached {fname}")
            else:
                data = get(BASE + fname)
                if data[:2] != b"PK":
                    print(f"FY{fy}: {fname} not a workbook, skip"); continue
                open(path, "wb").write(data)
                time.sleep(6)
            parsed = parse_fy(path, fy)
            for key, arr in parsed.items():
                series[key]["fy"][str(fy)] = arr
                for i, v in enumerate(arr):
                    if v is not None:
                        mm = ((i + 3) % 12) + 1
                        yy = fy + (1 if mm <= 3 else 0)
                        tag = f"{yy}-{mm:02d}"
                        if latest_month is None or tag > latest_month:
                            latest_month = tag
            try:
                players_raw.update(parse_players(path))
            except Exception as e:
                print(f"FY{fy}: players parse failed: {e}")
            print(f"FY{fy}: parsed {fname}  (foreign Apr={series['foreign']['fy'][str(fy)][0]})")
            time.sleep(6)
        except Exception as e:
            print(f"FY{fy}: FAILED {fname}: {e}")

    players = build_players(players_raw) if players_raw else None

    doc = {
        "meta": {
            "as_of": latest_month,
            "generated_at": datetime.datetime.utcnow().replace(microsecond=0).isoformat()+"Z",
            "unit": "JPY tn",
            "source": "JSDA 公社債投資家別売買高 (国債投資家別売買高) — outright (一般) net, super-long JGB >10Y",
            "note": "Net buying = purchases − sales, outright (ex-repo). Positive = net buying of super-long (>10Y) JGBs. One line per fiscal year (Apr–Mar).",
            "source_url": "https://www.jsda.or.jp/shiryoshitsu/toukei/toushika/index.html",
        },
        "months": MONTHS,
        "series": series,
        "players": players,
    }
    for dest in (os.path.join(HERE, "superlong_data.json"),
                 os.path.join(HERE, "streamlit_app", "superlong_data.json")):
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        json.dump(doc, open(dest, "w"), ensure_ascii=False, indent=1)
    print("as_of:", latest_month, "-> wrote superlong_data.json (root + streamlit_app)")

if __name__ == "__main__":
    main()
