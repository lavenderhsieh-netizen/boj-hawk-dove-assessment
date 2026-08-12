#!/usr/bin/env python3.12
"""JGB & T-Bills holdings by holder type — BOJ Flow of Funds (Financial Assets, Preliminary).

Source: https://www.boj.or.jp/en/statistics/sj/sjpre.xlsx ("Preliminary Figures", latest
quarter only — BOJ does not keep a stable URL for past quarters, so this script only ever
pulls the CURRENT quarter; re-run each quarter to roll the snapshot forward).

Item "Db" = "Central government securities and FILP bonds" (BOJ's own definition of
"Government bonds and treasury bills" = Da + Db; Da "Treasury discount bills" is ~immaterial
for most holder sectors and omitted from the by-holder breakdown here for simplicity — it's
carried in the "Other" residual).

Sheet map (sjpre.xlsx, this vintage):
  19 = All Sectors (1) — Financial institutions / Nonfinancial corp / General government(central)
  20 = All Sectors (2) — Local govt / Social security funds (incl. public pensions) / Households / PNPISH / Overseas / Total
  21 = Financial Institutions — Central bank(BOJ) / Depository corporations / Securities investment trusts
  25 = Insurance and Pension Funds (1) — Insurance (life/nonlife/mutual aid)
  26 = Insurance and Pension Funds (2) — Pension funds (corporate/other) + Pension total (=142+331)
"""
import json
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
URL = "https://www.boj.or.jp/en/statistics/sj/sjpre.xlsx"
XLSX = HERE / "source" / "flow_of_funds" / "sjpre.xlsx"


def download():
    import urllib.request
    XLSX.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        XLSX.write_bytes(r.read())


def find_row(ws, label, max_row=71):
    for r in range(1, max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, max_col=24))[0]
        for c in row:
            if c.value == label:
                return r
    raise ValueError(f"label {label!r} not found")


def cell(ws, row, col_letter):
    v = ws[f"{col_letter}{row}"].value
    return v if isinstance(v, (int, float)) else 0


def extract(path=None):
    wb = openpyxl.load_workbook(path or XLSX, data_only=True)

    # As-of date, e.g. "End of March 2026(Preliminary)"
    as_of = None
    ws19 = wb["19"]
    for r in range(1, 8):
        for c in list(ws19.iter_rows(min_row=r, max_row=r, max_col=24))[0]:
            if isinstance(c.value, str) and "End of" in c.value:
                as_of = c.value.replace("(Preliminary)", "").strip()
    if not as_of:
        as_of = "latest quarter"

    db19 = find_row(ws19, " Central government securities and FILP bonds")
    ws20 = wb["20"]
    db20 = find_row(ws20, "Db")  # code column is the marker on sheet 20, not the label
    ws21 = wb["21"]
    db21 = find_row(ws21, " Central government securities and FILP bonds")
    ws25 = wb["25"]
    db25 = find_row(ws25, " Central government securities and FILP bonds")
    ws26 = wb["26"]
    db26 = find_row(ws26, "Db")

    # units: 100 million yen (億円). Divide by 1e4 to get ¥tn.
    def tn(v):
        return round(v / 1e4, 2)

    total_all = tn(cell(ws20, db20, "O"))  # Total, all sectors, assets side

    boj = tn(cell(ws21, db21, "I"))            # Central bank
    banks = tn(cell(ws21, db21, "K"))           # Depository corporations
    inv_trusts = tn(cell(ws21, db21, "M"))      # Securities investment trusts
    insurance = tn(cell(ws25, db25, "G"))       # Insurance (life+nonlife+mutual aid)
    pension_private = tn(cell(ws26, db26, "C")) # Pension funds (corporate + other)
    public_pensions = tn(cell(ws20, db20, "G")) # Of which: public pensions (within social security funds)
    social_security_other = tn(cell(ws20, db20, "E")) - public_pensions  # residual social security funds ex-public-pension
    overseas = tn(cell(ws20, db20, "M"))        # Overseas (foreign investors)
    households = tn(cell(ws20, db20, "I"))      # Households
    local_govt = tn(cell(ws20, db20, "C"))      # Local governments
    central_govt = tn(cell(ws19, db19, "Q"))    # Central government (own holdings, small)
    nonfin_corp = tn(cell(ws19, db19, "I"))     # Nonfinancial corporations (private+public)
    pnpish = tn(cell(ws20, db20, "K"))          # Private nonprofit institutions serving households

    named = {
        "Bank of Japan": boj,
        "Banks": banks,
        "Insurance companies": insurance,
        "Public pensions (incl. GPIF)": public_pensions,
        "Pension funds (corporate/other)": pension_private,
        "Investment trusts": inv_trusts,
        "Overseas (foreign investors)": overseas,
        "Households": households,
    }
    other = round(total_all - sum(named.values()) - local_govt - central_govt - nonfin_corp - pnpish - social_security_other, 2)
    other = round(other + local_govt + central_govt + nonfin_corp + pnpish + social_security_other, 2)
    named["Other (nonfin. corp, gov't, other financial intermediaries)"] = other

    holders = [{"name": k, "jpy_tn": v, "pct": round(v / total_all * 100, 1)} for k, v in named.items()]
    holders.sort(key=lambda x: -x["jpy_tn"])

    out = {
        "as_of": as_of,
        "total_jpy_tn": total_all,
        "holders": holders,
        "source": "BOJ Flow of Funds Accounts, Preliminary Figures (sjpre.xlsx) — 'Government bonds and treasury bills' item, series Db (Central government securities and FILP bonds), by holder sector, assets side.",
        "source_url": URL,
    }
    return out


def main():
    try:
        download()
    except Exception as e:
        print(f"download failed ({e}); using existing local copy if present", file=sys.stderr)
        if not XLSX.exists():
            raise
    data = extract()
    out_path = HERE / "jgb_holders.json"
    out_path.write_text(json.dumps(data, indent=2))
    print(f"wrote {out_path} — as of {data['as_of']}, total {data['total_jpy_tn']} tn")
    for h in data["holders"]:
        print(f"  {h['name']:<45} {h['jpy_tn']:>8.1f} tn  ({h['pct']}%)")


if __name__ == "__main__":
    main()
