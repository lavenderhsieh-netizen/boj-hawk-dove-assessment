#!/usr/bin/env python3
"""Fetch public market, inflation and news data for the BOJ hawk-dove dashboard.

Sources:
  - MOF: daily JGB yield curve (jgbcme.csv, + historical backfill)
  - FRED: USD/JPY (DEXJPUS), Nikkei 225 (NIKKEI225), US 10Y (DGS10),
          uncollateralized overnight call rate (IRSTCI01JPM156N)
  - BOJ: measures of underlying inflation (cpirev.xlsx)
  - Google News RSS: BOJ / Japan macro headlines

Each source is fetched and validated independently. If a source fails,
the previous good section in market_data.json is kept and the source is
marked "stale" — the dashboard keeps showing the most recent correct data.

Usage: /usr/local/bin/python3 fetch_market.py
Requires: requests, openpyxl. FRED_API_KEY env var.
"""

import csv
import io
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime

import requests

REPO = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(REPO, "market_data.json")
UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"}
FRED_KEY = os.environ.get("FRED_API_KEY", "")

MOF_CURRENT = "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/jgbcme.csv"
MOF_HISTORICAL = "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/historical/jgbcme_all.csv"
BOJ_CPI = "https://www.boj.or.jp/en/research/research_data/cpi/cpirev.xlsx"
BOJ_GAP = "https://www.boj.or.jp/en/research/research_data/gap/gap.xlsx"

HISTORY_DAYS = 400          # trading days kept for line charts
CURVE_HISTORY_DAYS = 90     # daily full curves kept for curve comparison


def get(url, **kw):
    r = requests.get(url, headers=UA, timeout=45, **kw)
    r.raise_for_status()
    return r


def fred_series(series_id, days_back=550):
    start = (datetime.now(timezone.utc) - timedelta(days=days_back)).strftime("%Y-%m-%d")
    url = (
        "https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&api_key={FRED_KEY}&file_type=json"
        f"&observation_start={start}"
    )
    data = get(url).json()
    hist = []
    for o in data.get("observations", []):
        if o["value"] in (".", "", None):
            continue
        hist.append({"date": o["date"], "value": float(o["value"])})
    return hist


def check(cond, msg):
    if not cond:
        raise ValueError(msg)


# ── Source fetchers ──────────────────────────────────────────────────────────

def parse_mof_csv(text):
    """Parse a MOF jgbcme CSV into [{date: 'YYYY-MM-DD', yields: {tenor: float}}]."""
    rows = list(csv.reader(io.StringIO(text)))
    tenors, out = None, []
    for row in rows:
        if not row or not row[0].strip():
            continue
        if row[0].strip() == "Date":
            tenors = [t.strip() for t in row[1:] if t.strip()]
            continue
        if tenors is None or not re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", row[0].strip()):
            continue
        d = datetime.strptime(row[0].strip(), "%Y/%m/%d").strftime("%Y-%m-%d")
        yields = {}
        for tenor, val in zip(tenors, row[1:]):
            val = val.strip()
            if val and val != "-":
                try:
                    yields[tenor] = float(val)
                except ValueError:
                    pass
        if yields:
            out.append({"date": d, "yields": yields})
    return out


def fetch_jgb(prev):
    curves = parse_mof_csv(get(MOF_CURRENT).text)

    # backfill long history once (or if history is thin)
    prev_hist = (prev or {}).get("history", [])
    if len(prev_hist) < 100:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=720)).strftime("%Y-%m-%d")
        all_curves = parse_mof_csv(get(MOF_HISTORICAL).text)
        curves = [c for c in all_curves if c["date"] >= cutoff] + curves
    else:
        # Always backfill full curves so 1-month curve comparison works
        # (MOF_CURRENT only has a few days; full-curve history needs 28+ days)
        target_full = (datetime.now(timezone.utc) - timedelta(days=35)).strftime("%Y-%m-%d")
        if not any(c["date"] <= target_full and len(c["yields"]) >= 10 for c in curves):
            cutoff = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
            hist_curves = parse_mof_csv(get(MOF_HISTORICAL).text)
            extra = [c for c in hist_curves if c["date"] >= cutoff and c["date"] < curves[0]["date"]]
            curves = extra + curves

    # merge with previous history (keyed by date)
    by_date = {}
    for h in prev_hist:
        by_date[h["date"]] = {"date": h["date"], "yields": {
            "2Y": h.get("y2"), "10Y": h.get("y10"), "30Y": h.get("y30")}}
    for c in curves:
        by_date[c["date"]] = c

    dates = sorted(by_date)
    check(len(dates) > 0, "no JGB data")
    latest = by_date[dates[-1]]
    y10 = latest["yields"].get("10Y")
    check(y10 is not None and -1 < y10 < 10, f"JGB 10Y out of range: {y10}")
    check(len(latest["yields"]) >= 10, "JGB curve too sparse")
    check(latest["date"] >= (datetime.now(timezone.utc) - timedelta(days=21)).strftime("%Y-%m-%d"),
          f"JGB data too old: {latest['date']}")

    hist = [{"date": d,
             "y2": by_date[d]["yields"].get("2Y"),
             "y10": by_date[d]["yields"].get("10Y"),
             "y30": by_date[d]["yields"].get("30Y")}
            for d in dates][-HISTORY_DAYS:]

    # curve ~1 month ago: latest date <= latest-28d, else earliest kept full curve
    target = (datetime.strptime(latest["date"], "%Y-%m-%d") - timedelta(days=28)).strftime("%Y-%m-%d")
    full = [d for d in dates if len(by_date[d]["yields"]) >= 10]
    older = [d for d in full if d <= target]
    prev_curve_date = older[-1] if older else full[0]
    prev_curve = by_date[prev_curve_date]

    # curve 1 business day before latest
    full_before = [d for d in full if d < latest["date"]]
    curve_1d_date = full_before[-1] if full_before else full[0]
    curve_1d = by_date[curve_1d_date]

    tenor_order = ["1Y", "2Y", "3Y", "4Y", "5Y", "6Y", "7Y", "8Y", "9Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y"]
    tenors = [t for t in tenor_order if t in latest["yields"]]
    return {
        "curve": {"date": latest["date"], "tenors": tenors,
                  "yields": [latest["yields"][t] for t in tenors]},
        "curve_1d": {"date": curve_1d["date"], "tenors": tenors,
                     "yields": [curve_1d["yields"].get(t) for t in tenors]},
        "curve_prev": {"date": prev_curve["date"], "tenors": tenors,
                       "yields": [prev_curve["yields"].get(t) for t in tenors]},
        "history": hist,
    }


def fetch_fred_single(series_id, lo, hi, name, days_back=550):
    hist = fred_series(series_id, days_back)
    check(len(hist) > 10, f"{name}: too few observations")
    latest = hist[-1]
    check(lo < latest["value"] < hi, f"{name} out of range: {latest['value']}")
    check(latest["date"] >= (datetime.now(timezone.utc) - timedelta(days=45)).strftime("%Y-%m-%d"),
          f"{name} too old: {latest['date']}")
    return {"history": hist}


def fetch_call_rate(prev):
    hist = fred_series("IRSTCI01JPM156N", 800)
    check(len(hist) > 6, "call rate: too few observations")
    latest = hist[-1]
    check(-1 < latest["value"] < 5, f"call rate out of range: {latest['value']}")
    check(latest["date"] >= (datetime.now(timezone.utc) - timedelta(days=100)).strftime("%Y-%m-%d"),
          f"call rate too old: {latest['date']}")
    return {"history": hist}


def fetch_inflation(prev):
    import openpyxl
    r = get(BOJ_CPI)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    months, ex_fresh, ex_fresh_energy, ex_food_energy = [], [], [], []
    for row in ws.iter_rows(min_row=6, values_only=True):
        d = row[0]
        if not isinstance(d, datetime) or d.year < 2019:
            continue
        b, e, h = row[1], row[4], row[7]
        if b is None and e is None and h is None:
            continue
        months.append(d.strftime("%Y-%m"))
        ex_fresh.append(round(b, 2) if isinstance(b, (int, float)) else None)
        ex_fresh_energy.append(round(e, 2) if isinstance(e, (int, float)) else None)
        ex_food_energy.append(round(h, 2) if isinstance(h, (int, float)) else None)
    check(len(months) > 12, "inflation: too few months")
    last_month = datetime.strptime(months[-1], "%Y-%m")
    check((datetime.now() - last_month).days < 120, f"inflation too old: {months[-1]}")
    vals = [v for v in ex_fresh if v is not None]
    check(all(-5 < v < 15 for v in vals), "inflation out of range")
    return {"months": months, "ex_fresh": ex_fresh,
            "ex_fresh_energy": ex_fresh_energy, "ex_food_energy": ex_food_energy}


def fetch_potential_growth(prev):
    import openpyxl
    r = get(BOJ_GAP)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    # data1: quarterly output gap
    ws1 = wb["data1"]
    gap_rows = []
    for row in ws1.iter_rows(min_row=6, values_only=True):
        period, val = row[0], row[1]
        if not isinstance(period, str) or not isinstance(val, (int, float)):
            continue
        gap_rows.append({"date": period, "value": round(val, 3)})
    gap_rows = gap_rows[-24:]  # last 6 years of quarters
    check(len(gap_rows) > 4, "output gap: too few rows")

    # data2: semi-annual potential growth rate with decomposition
    ws2 = wb["data2"]
    pg_rows = []
    for row in ws2.iter_rows(min_row=6, values_only=True):
        period, rate, tfp, capital, hours, employed = row[0], row[1], row[2], row[3], row[4], row[5]
        if not isinstance(period, str) or not isinstance(rate, (int, float)):
            continue
        label = period.split(":")[0].strip()  # e.g. "2025.2"
        pg_rows.append({
            "date": label,
            "value": round(rate, 3),
            "tfp": round(tfp, 3) if isinstance(tfp, (int, float)) else None,
            "capital": round(capital, 3) if isinstance(capital, (int, float)) else None,
            "hours": round(hours, 3) if isinstance(hours, (int, float)) else None,
            "employed": round(employed, 3) if isinstance(employed, (int, float)) else None,
        })
    pg_rows = pg_rows[-20:]  # last ~10 years semi-annual
    check(len(pg_rows) > 4, "potential growth: too few rows")

    latest_gap = gap_rows[-1]["value"] if gap_rows else None
    latest_pot = pg_rows[-1]["value"] if pg_rows else None
    latest_date = pg_rows[-1]["date"] if pg_rows else None
    check(latest_pot is not None, "potential growth: no latest value")

    return {
        "output_gap": gap_rows,
        "potential_rate": pg_rows,
        "latest_gap": latest_gap,
        "latest_pot": latest_pot,
        "latest_date": latest_date,
    }


NEWS_QUERIES = [
    # Source-specific queries — highest-quality outlets first
    'site:reuters.com Japan BOJ OR yen OR JGB when:3d',
    'site:bloomberg.com Japan BOJ OR yen OR JGB when:3d',
    'site:ft.com Japan BOJ OR yen OR fiscal when:3d',
    'site:wsj.com Japan BOJ OR yen OR JGB when:3d',
    'site:asia.nikkei.com Japan BOJ OR yen OR JGB when:3d',
    # Broader topic queries as fallback
    '\"Bank of Japan\" OR \"BOJ\" Ueda when:3d',
    '\"JGB\" OR \"Japan government bond\" yields when:3d',
    '\"Japan inflation\" OR \"Japan CPI\" when:3d',
    'Takaichi Japan economy OR fiscal OR BOJ when:3d',
    'Katayama Japan yen OR intervention when:3d',
    '\"Japan defense spending\" OR \"Japan rearmament\" when:3d',
    '\"BOJ rate hike\" OR \"Bank of Japan rate\" when:3d',
    '\"Japan yen\" intervention OR \"40-year\" when:3d',
]

TAG_RULES = [
    ("jgb-yield",       ["jgb", "government bond", "bond yield", "10-year jgb", "30-year jgb",
                         "yield curve", "jgb auction", "coupon", "bond market", "yields rise",
                         "yields fall", "benchmark yield"]),
    ("monetary-policy", ["bank of japan", "boj", "rate hike", "policy rate", "monetary policy",
                         "rate decision", "quantitative easing", "qe", "rate cut", "underlying inflation",
                         "price stability"]),
    ("fiscal",          ["fiscal", "budget", "national debt", "debt-to-gdp", "deficit", "fiscal spending",
                         "investment blueprint", "economic blueprint", "tax revenue", "gdp growth target"]),
    ("fx-yen",          ["yen", "fx intervention", "forex", "currency intervention", "exchange rate",
                         "dollar/yen", "usd/jpy", "40-year low", "yen weakness", "yen slide"]),
    ("takaichi",        ["takaichi"]),
    ("katayama",        ["katayama"]),
    ("ueda",            ["ueda"]),
    ("defense",         ["defense spending", "defence spending", "military spending", "rearmament",
                         "defense budget", "2% gdp", "security spending"]),
]

EXCLUDE_TITLE_WORDS = [
    "ukraine", "russia", "missile assault", "crypto briefing", "bitcoin", "nba",
    "world cup", "soccer", "football match", "spacex", "south korea", "taiwan ai",
]

TRUSTED_SOURCES = {
    "bloomberg", "reuters", "financial times", "ft.com",
    "wall street journal", "wsj", "nikkei", "nikkei asia",
    "cnbc", "the economist", "barron's", "marketwatch",
    "south china morning post", "scmp", "japan times",
    "associated press", "ap", "new york times",
}


def is_trusted(source):
    sl = source.lower()
    return any(t in sl for t in TRUSTED_SOURCES)


REQUIRED_WORDS = [
    "japan", "boj", "jgb", "yen", "nikkei", "takaichi", "katayama", "ueda",
    "bank of japan", "japanese", "tokyo",
]


def tag_item(title):
    tl = title.lower()
    tags = []
    for tag, keywords in TAG_RULES:
        if any(kw in tl for kw in keywords):
            tags.append(tag)
    return tags


def is_relevant(title):
    tl = title.lower()
    if any(ex in tl for ex in EXCLUDE_TITLE_WORDS):
        return False
    # Tier 1: specific monetary/financial keyword — always relevant
    FINANCIAL = [
        "boj", "bank of japan", "jgb", "yen", "takaichi", "katayama", "ueda",
        "rate hike", "rate cut", "monetary policy", "policy rate", "inflation",
        "bond yield", "fiscal", "government bond", "currency intervention",
        "fx intervention", "quantitative", "nikkei 225", "10-year", "30-year",
        "40-year", "yield curve", "bond market", "interest rate",
    ]
    if any(kw in tl for kw in FINANCIAL):
        return True
    # Tier 2: Japan/Tokyo + at least one economy/market word
    JAPAN_WORDS = ["japan", "japanese", "tokyo"]
    ECONOMY_WORDS = ["economy", "economic", "finance", "financial", "market",
                     "rate", "bond", "currency", "gdp", "budget", "debt",
                     "trade", "export", "import", "wage", "price", "growth"]
    if any(j in tl for j in JAPAN_WORDS) and any(e in tl for e in ECONOMY_WORDS):
        return True
    return False


def fetch_news(prev):
    items, seen = [], set()
    for q in NEWS_QUERIES:
        url = ("https://news.google.com/rss/search?q=" + requests.utils.quote(q)
               + "&hl=en-US&gl=US&ceid=US:en")
        try:
            root = ET.fromstring(get(url).content)
        except Exception:
            continue
        for it in root.iter("item"):
            title = (it.findtext("title") or "").strip()
            link = (it.findtext("link") or "").strip()
            pub = (it.findtext("pubDate") or "").strip()
            src = (it.findtext("source") or "").strip()
            if not title or not link:
                continue
            base = title.rsplit(" - ", 1)[0].strip().lower()
            if base in seen:
                continue
            seen.add(base)
            try:
                dt = parsedate_to_datetime(pub)
                pub_iso = dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            except Exception:
                pub_iso = None
            if not src and " - " in title:
                title, src = title.rsplit(" - ", 1)
            title = title.strip()
            if not is_relevant(title):
                continue
            tags = tag_item(title)
            items.append({"title": title, "link": link,
                          "source": src or "Google News", "published": pub_iso,
                          "tags": tags})
    items = [i for i in items if i["published"]]
    # Drop articles older than yesterday midnight UTC — keeps today + yesterday, excludes anything older
    cutoff = (datetime.now(timezone.utc) - timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    ).strftime("%Y-%m-%dT%H:%M:%SZ")
    items = [i for i in items if i["published"] >= cutoff]
    items.sort(key=lambda i: i["published"], reverse=True)
    # Prefer trusted sources; fill with others only if fewer than 10 trusted items
    trusted = [i for i in items if is_trusted(i["source"])]
    others  = [i for i in items if not is_trusted(i["source"])]
    if len(trusted) >= 10:
        items = trusted[:30]
    else:
        items = (trusted + others)[:30]
    check(len(items) >= 3, f"news: only {len(items)} items")
    return items


# ── Orchestration with last-good fallback ────────────────────────────────────

SOURCES = {
    "jgb":              lambda prev: fetch_jgb(prev),
    "fx":               lambda prev: fetch_fred_single("DEXJPUS", 50, 300, "USD/JPY"),
    "nikkei":           lambda prev: fetch_fred_single("NIKKEI225", 5000, 200000, "Nikkei"),
    "us10y":            lambda prev: fetch_fred_single("DGS10", 0, 20, "US 10Y"),
    "call_rate":        fetch_call_rate,
    "inflation":        fetch_inflation,
    "potential_growth": fetch_potential_growth,
    "news":             fetch_news,
}


def latest_date_of(key, section):
    try:
        if key == "jgb":
            return section["curve"]["date"]
        if key == "inflation":
            return section["months"][-1]
        if key == "potential_growth":
            return section["latest_date"]
        if key == "news":
            return section[0]["published"][:10]
        return section["history"][-1]["date"]
    except Exception:
        return None


def main():
    if not FRED_KEY:
        print("WARNING: FRED_API_KEY not set — FRED sources will fail", file=sys.stderr)

    old = {}
    if os.path.exists(OUT_PATH):
        try:
            with open(OUT_PATH) as f:
                old = json.load(f)
        except Exception:
            old = {}

    out = {"updated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "sources": {}}
    old_sources = old.get("sources", {})
    failures = 0

    for key, fn in SOURCES.items():
        prev_section = old.get(key)
        try:
            section = fn(prev_section)
            out[key] = section
            out["sources"][key] = {"status": "ok", "error": None,
                                   "latest": latest_date_of(key, section),
                                   "fetched_at": out["updated_utc"]}
            print(f"[ok]    {key}: latest={out['sources'][key]['latest']}")
        except Exception as e:
            failures += 1
            if prev_section is not None:
                out[key] = prev_section
                meta = dict(old_sources.get(key, {}))
                meta.update({"status": "stale", "error": str(e)[:300]})
                meta.setdefault("latest", latest_date_of(key, prev_section))
                out["sources"][key] = meta
                print(f"[stale] {key}: kept last good ({meta.get('latest')}) — {e}", file=sys.stderr)
            else:
                out["sources"][key] = {"status": "missing", "error": str(e)[:300],
                                       "latest": None, "fetched_at": None}
                print(f"[fail]  {key}: no previous data to fall back on — {e}", file=sys.stderr)

    tmp = OUT_PATH + ".tmp"
    with open(tmp, "w") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, OUT_PATH)
    size_kb = os.path.getsize(OUT_PATH) / 1024
    print(f"wrote {OUT_PATH} ({size_kb:.0f} KB), {failures} source failure(s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
